from typing import Tuple, List, Union
import openpyxl


def write_to_sheet(data: List[Union[tuple, str]], output_file: str):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    for row, value in enumerate(data, start=1):
        if isinstance(value, tuple):
            for col, v in enumerate(value, start=1):
                sheet.cell(row=row, column=col).value = v
        else:
            sheet.cell(row=row, column=1).value = value

    workbook.save(output_file)
