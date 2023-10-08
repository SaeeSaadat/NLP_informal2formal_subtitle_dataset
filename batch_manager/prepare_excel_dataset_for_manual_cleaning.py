import re

import openpyxl
from hazm import SentenceTokenizer
from xlsxs import xlsx_writer, xlsx_reader
from cleanup.data_cleaner import split_sentences_by_punkt


def divide_excel_file(xlsx_file: str, section_count: int):
    sheet = openpyxl.load_workbook(xlsx_file).active
    num_of_rows = sheet.max_row
    num_of_sections = num_of_rows // section_count
    for i in range(num_of_sections+1):
        start_row = i * section_count + 1
        end_row = (i + 1) * section_count

        if i == section_count - 1:
            end_row = num_of_rows
        new_sheet = openpyxl.Workbook().active
        for row in sheet.iter_rows(min_row=start_row, max_row=end_row):
            new_sheet.append([cell.value for cell in row])
        new_sheet.title = f'section_{i}'
        new_sheet.parent.save(f'../resources/handmade_dataset/sections/dirty_section_{i}.xlsx')


def prepare_long_dataset(input_file: str):
    sentence_tokenizer = SentenceTokenizer()
    sheet = openpyxl.load_workbook(input_file).active
    output_data = []

    for row in sheet.iter_rows():
        informal = row[0].value
        formal = row[1].value
        sentences = sentence_tokenizer.tokenize(informal)
        is_first = True
        for sentence in sentences:
            for s in split_sentences_by_punkt(sentence):
                if is_first:
                    output_data.append((s, s, formal))
                    is_first = False
                else:
                    output_data.append((s, s, ''))
        output_data.append(('', '', ''))
    xlsx_writer.write_to_sheet(output_data, f'../resources/handmade_dataset/long_600_digikala_divided.xlsx')


def merge_files(file1: str, file2: str):
    data1 = xlsx_reader.read_from_xlsx(file1)
    print(data1)
    data2 = xlsx_reader.read_from_xlsx(file2)
    data2 = [(row[1], row[0], '') for row in data2]

    xlsx_writer.write_to_sheet(data1 + data2, '../resources/handmade_dataset/dirty_800_merged.xlsx')


if __name__ == '__main__':
    # prepare_long_dataset('../resources/handmade_dataset/long_600_digikala.xlsx')
    # merge_files('../resources/handmade_dataset/long_600_digikala_divided.xlsx',
    #             '../resources/handmade_dataset/200-ninisite.xlsx')
    divide_excel_file('../resources/handmade_dataset/dirty_800_merged.xlsx', 700)
